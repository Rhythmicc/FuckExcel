import openpyxl
import os

user_root = os.path.expanduser('~')


class FuckExcel:
    def __init__(self, excel_path):
        self.file_path = os.path.abspath(excel_path.replace('~', user_root))
        if os.path.exists(self.file_path):
            self.excel = openpyxl.load_workbook(self.file_path)
            self.sheet = self.excel.get_sheet_by_name(self.excel.get_sheet_names()[0])
        else:
            self.excel = openpyxl.Workbook()
            self.sheet = self.excel.active
            self.sheet.title = 'Sheet1'

    def set_work_sheet(self, sheet_name):
        self.sheet = self.excel.get_sheet_by_name(sheet_name)

    def sheet_size(self):
        return self.sheet.max_row, self.sheet.max_column

    def get_sheet_names(self):
        return self.excel.get_sheet_names()

    def is_empty_row(self, row):
        for i in range(1, self.sheet.max_column + 1):
            if self.sheet.cell(row, i).value:
                return False
        return True

    def is_empty_column(self, column):
        for i in range(1, self.sheet.max_column + 1):
            if self.sheet.cell(i, column).value:
                return False
        return True

    def __setitem__(self, key, value):
        if isinstance(key, tuple):
            if isinstance(key[0], int):
                if isinstance(key[1], int):
                    self.sheet.cell(key[0], key[1]).value = value
                else:
                    for j in range(key[1].start if key[1].start else 1,
                                   (key[1].stop if key[1].stop else self.sheet_size()[1]) + 1,
                                   key[1].step if key[1].step else 1):
                        self.sheet.cell(key[0], j).value = value
            else:
                for i in range(key[0].start if key[0].start else 1,
                               (key[0].stop if key[0].stop else self.sheet_size()[0]) + 1,
                               key[0].step if key[0].step else 1):
                    if isinstance(key[1], int):
                        self.sheet.cell(i, key[1]).value = value
                    else:
                        for j in range(key[1].start if key[1].start else 1,
                                       (key[1].stop if key[1].stop else self.sheet_size()[1]) + 1,
                                       key[1].step if key[1].step else 1):
                            self.sheet.cell(i, j).value = value
        else:
            raise IndexError('Index must be a tuple like (1, 1)')

    def __getitem__(self, item):
        if isinstance(item, tuple):
            item = list(item)
            if isinstance(item[0], int):
                if item[0] < 0:
                    item[0] += self.sheet_size()[0] + 1
                if isinstance(item[1], int):
                    if item[1] < 0:
                        item[1] += self.sheet_size()[1] + 1
                    return self.sheet.cell(item[0], item[1]).value
                else:
                    start, stop = item[1].start if not item[1].start or item[1].start > 0 else self.sheet_size()[1] + 1 + item[1].start, item[1].stop if not item[1].stop or item[1].stop > 0 else self.sheet_size()[1] + 1 + item[1].stop
                    return [self.sheet.cell(item[0], i).value for i in range(start if start else 1, stop if stop else self.sheet_size()[1] + 1, item[1].step if item[1].step else 1)]
            else:
                start0, stop0 = item[0].start if not item[0].start or item[0].start > 0 else self.sheet_size()[0] + 1 + item[0].start, item[0].stop if not item[0].stop or item[0].stop > 0 else self.sheet_size()[0] + 1 + item[0].stop
                if isinstance(item[1], int):
                    if item[1] < 0:
                        item[1] += self.sheet_size()[1] + 1
                    return [self.sheet.cell(i, item[1]).value for i in range(start0 if start0 else 1, stop0 if stop0 else self.sheet_size()[0] + 1, item[0].step if item[0].step else 1)]
                else:
                    start1, stop1 = item[1].start if not item[1].start or item[1].start > 0 else self.sheet_size()[1] + 1 + item[1].start, item[1].stop if not item[1].stop or item[1].stop > 0 else self.sheet_size()[1] + 1 + item[1].stop
                    return [[self.sheet.cell(i, j).value for i in range(start0 if start0 else 1, stop0 if stop0 else self.sheet_size()[0] + 1, item[0].step if item[0].step else 1)]
                            for j in range(start1 if start1 else 1, stop1 if stop1 else self.sheet_size()[1] + 1, item[1].step if item[1].step else 1)]
        else:
            raise IndexError('Index must be a tuple like (1, 1)')

    def append_row(self, vals):
        sz = self.sheet_size()
        for i, val in enumerate(vals):
            self[sz[0], i+1] = val

    def append_column(self, vals):
        sz = self.sheet_size()
        for i, val in enumerate(vals):
            self[i+1, sz[1]] = val

    def push_back_row(self, row, value):
        self[row, self.sheet_size()[1]] = value

    def push_back_colunm(self, column, value):
        self[self.sheet_size()[0], column] = value

    def delete_row(self, row_num):
        sz = self.sheet_size()
        for i in range(row_num, sz[0]):
            for j in range(1, sz[1]+1):
                self[i, j] = self[i+1, j]
        for i in range(1, sz[1]+1):
            self[sz[0], i] = None

    def delete_column(self, column_num):
        sz = self.sheet_size()
        for i in range(column_num, sz[1]):
            for j in range(1, sz[0] + 1):
                self[j, i] = self[j, i+1]
        for i in range(1, sz[0] + 1):
            self[i, sz[1]] = None

    def save(self, path=None):
        self.excel.save(self.file_path if not path else path)

    def clear_all(self):
        sz = self.sheet_size()
        for i in range(1, sz[0]+1):
            for j in range(1, sz[1]+1):
                self[i, j] = None
